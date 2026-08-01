#!/usr/bin/env python3
"""
Resync the Studio Hub revenue dashboard (analytics.html) from the Excel Deal Tracker.

Every figure is derived from the raw deal rows on the "2026 ACTUAL" sheet — one row
per deal — NOT from the "Analytics Dashboard" summary tab. That summary tab keeps
hand-typed lists of clients, industries and services, so anything new silently drops
out of it; reading the deal rows means a new client shows up the moment their first
deal is entered, with no maintenance.

The payload is encrypted with the dashboard passcode (AES-GCM / PBKDF2-SHA256,
matching the WebCrypto code in the page) and written over the `const BLOB="..."` line.

The passcode is never stored in this repo. It is read from, in order:
  1. $HUB_KEY
  2. macOS Keychain:  security find-generic-password -s anomaly-hub-key -w

Store it once with:
  security add-generic-password -a "$USER" -s anomaly-hub-key -w

Usage:
  python3 scripts/sync_revenue.py              # write analytics.html
  python3 scripts/sync_revenue.py --dry-run    # print the payload, touch nothing
  python3 scripts/sync_revenue.py --push       # write, commit analytics.html, push
"""

import argparse
import base64
import datetime as dt
import hashlib
import json
import os
import re
import subprocess
import sys
import warnings

warnings.filterwarnings("ignore")

# ── config ────────────────────────────────────────────────────────────────────
REPO = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PAGE = os.path.join(REPO, "analytics.html")
XLSX = "/Users/danielpan/Dropbox/Anomaly Creative LLC/Deal Tracker.xlsx"
DEALS = "2026 ACTUAL"          # one row per deal — the source of truth
SUMMARY = "Analytics Dashboard"  # hand-maintained; used only as a cross-check

# Columns on the DEALS sheet (0-based).
CLIENT, MONTH, AMOUNT, INDUSTRY = 0, 2, 3, 4
SERVICES = range(5, 10)        # Service 1 … Service 5
COLLECTED = 10                 # "Payment Collected"

# Yearly collection goals for the "projected vs goals" chart.
GOALS = {"conservative": 800_000, "moderate": 1_000_000, "aggressive": 1_200_000}

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

PBKDF2_ITERS = 150_000  # must match analytics.html


# ── passcode ──────────────────────────────────────────────────────────────────
def get_passcode() -> str:
    key = os.environ.get("HUB_KEY")
    if key:
        return key.strip()
    try:
        out = subprocess.run(
            ["security", "find-generic-password", "-s", "anomaly-hub-key", "-w"],
            capture_output=True, text=True, check=True,
        )
        return out.stdout.strip()
    except Exception:
        sys.exit(
            "No dashboard passcode found.\n"
            "  Store it once:  security add-generic-password -a \"$USER\" "
            "-s anomaly-hub-key -w\n"
            "  Or export HUB_KEY for a one-off run."
        )


def gate_hash_from_page(html: str) -> str:
    m = re.search(r'const GATE_HASH="([0-9a-f]{64})"', html)
    if not m:
        sys.exit("Could not find GATE_HASH in analytics.html.")
    return m.group(1)


# ── workbook ──────────────────────────────────────────────────────────────────
def txt(v):
    return str(v).strip() if v is not None else ""


def is_true(v):
    return v is True or txt(v).upper() == "TRUE"


def build_payload():
    import openpyxl
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    if DEALS not in wb.sheetnames:
        sys.exit(f'Sheet "{DEALS}" not found in {XLSX}')

    projected = [0] * 12
    collected = [0] * 12
    by_client, by_industry, by_service = {}, {}, {}
    skipped = 0

    for r in wb[DEALS].iter_rows(min_row=2, values_only=True):
        client = txt(r[CLIENT]) if len(r) > CLIENT else ""
        amount = r[AMOUNT] if len(r) > AMOUNT and isinstance(r[AMOUNT], (int, float)) else 0
        month = r[MONTH] if len(r) > MONTH else None
        if not client or not amount:
            continue
        if not isinstance(month, (dt.datetime, dt.date)):
            skipped += 1          # a deal with no usable date can't land in a month
            continue

        i = month.month - 1
        projected[i] += amount
        if is_true(r[COLLECTED] if len(r) > COLLECTED else None):
            collected[i] += amount

        by_client[client] = by_client.get(client, 0) + amount
        ind = txt(r[INDUSTRY]) if len(r) > INDUSTRY else ""
        if ind:
            by_industry[ind] = by_industry.get(ind, 0) + amount
        for c in SERVICES:
            s = txt(r[c]) if len(r) > c else ""
            if s:
                by_service.setdefault(s, set()).add(client)

    if not by_client:
        sys.exit(f'No usable deal rows found on "{DEALS}".')

    total_p, total_c = sum(projected), sum(collected)

    # complete months only — the month in progress would drag the run rate down
    today = dt.date.today()
    year = max((r[MONTH].year for r in wb[DEALS].iter_rows(min_row=2, values_only=True)
                if len(r) > MONTH and isinstance(r[MONTH], (dt.datetime, dt.date))),
               default=today.year)
    elapsed = 12 if year < today.year else max(1, min(12, today.month - 1))

    payload = {
        "asOf": dt.date.fromtimestamp(os.path.getmtime(XLSX)).strftime("%b %-d, %Y"),
        "elapsed": elapsed,
        "summary": {
            "projected": round(total_p),
            "collected": round(total_c),
            "outstanding": round(total_p - total_c),
            "clients": len(by_client),
        },
        "months": MONTHS,
        "projected": [round(v) for v in projected],
        "collected": [round(v) for v in collected],
        "industries": sorted(([k, round(v)] for k, v in by_industry.items()),
                             key=lambda x: -x[1]),
        "services": sorted(([k, len(v)] for k, v in by_service.items()),
                           key=lambda x: -x[1]),
        "clientRanking": sorted(([k, round(v)] for k, v in by_client.items()),
                                key=lambda x: -x[1]),
        "goals": GOALS,
    }

    # cross-check against the hand-maintained summary tab and say so if they diverge
    notes = []
    if skipped:
        notes.append(f"{skipped} deal row(s) had no usable date and were skipped")
    if SUMMARY in wb.sheetnames:
        ad = wb[SUMMARY]
        for label, got in (("Total Projected Revenue", total_p),
                           ("Total Collected Revenue", total_c)):
            for row in ad.iter_rows(max_col=2, values_only=True):
                if txt(row[0]) == label and isinstance(row[1], (int, float)):
                    if round(row[1]) != round(got):
                        notes.append(f"{label}: deals say ${got:,.0f}, "
                                     f"summary tab says ${row[1]:,.0f}")
                    break
    return payload, notes


# ── crypto (mirrors the WebCrypto code in analytics.html) ─────────────────────
def encrypt(payload: dict, passcode: str) -> str:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    salt, nonce = os.urandom(16), os.urandom(12)
    key = hashlib.pbkdf2_hmac("sha256", passcode.encode(), salt, PBKDF2_ITERS, 32)
    ct = AESGCM(key).encrypt(nonce, json.dumps(payload, separators=(",", ":")).encode(), None)
    return base64.b64encode(salt + nonce + ct).decode()


def decrypt(blob: str, passcode: str) -> dict:
    from cryptography.hazmat.primitives.ciphers.aead import AESGCM
    raw = base64.b64decode(blob)
    key = hashlib.pbkdf2_hmac("sha256", passcode.encode(), raw[:16], PBKDF2_ITERS, 32)
    return json.loads(AESGCM(key).decrypt(raw[16:28], raw[28:], None))


def previous_payload(html: str, passcode: str):
    """The payload currently published, so we can report what changed. None if unreadable."""
    m = re.search(r'const BLOB="([^"]*)"', html)
    try:
        return decrypt(m.group(1), passcode) if m else None
    except Exception:
        return None


def run_rate(p):
    win = min(3, p["elapsed"])
    avg = sum(p["projected"][p["elapsed"] - win:p["elapsed"]]) / win
    return avg, avg * 12, win


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="print the payload, write nothing")
    ap.add_argument("--push", action="store_true", help="commit analytics.html and push")
    args = ap.parse_args()

    payload, notes = build_payload()
    avg, rr, win = run_rate(payload)

    if args.dry_run:
        print(json.dumps(payload, indent=2))
        print(f"\nrun rate: mean of the last {win} complete months of projected "
              f"collections (${avg:,.0f}) x 12 = ${rr:,.0f}")
        for n in notes:
            print("NOTE:", n)
        return

    with open(PAGE, encoding="utf-8") as f:
        html = f.read()

    passcode = get_passcode()
    if hashlib.sha256(("anomaly-hub|" + passcode).encode()).hexdigest() != gate_hash_from_page(html):
        sys.exit("Passcode does not match the dashboard gate — refusing to write. "
                 "Nothing was changed.")

    # what's new since the version currently published
    prev = previous_payload(html, passcode)
    new_clients, gone_clients = [], []
    if prev:
        was = {c[0] for c in prev.get("clientRanking", [])}
        now = {c[0] for c in payload["clientRanking"]}
        new_clients = sorted(now - was, key=lambda n: -dict(payload["clientRanking"])[n])
        gone_clients = sorted(was - now)

    blob = encrypt(payload, passcode)
    if decrypt(blob, passcode) != payload:          # round-trip before we touch the page
        sys.exit("Encrypt/decrypt round-trip failed — refusing to write.")

    new_html, n = re.subn(r'const BLOB="[^"]*"', 'const BLOB="' + blob + '"', html, count=1)
    if n != 1:
        sys.exit("Could not locate the BLOB line in analytics.html.")
    with open(PAGE, "w", encoding="utf-8") as f:
        f.write(new_html)

    s = payload["summary"]
    print(f"analytics.html resynced — as of {payload['asOf']}: "
          f"${s['collected']:,} collected of ${s['projected']:,} projected, "
          f"{s['clients']} clients, {payload['elapsed']} months elapsed.")
    print(f"run rate: ${avg:,.0f} x 12 = ${rr:,.0f}")
    if new_clients:
        rank = dict(payload["clientRanking"])
        print(f"NEW CLIENTS ({len(new_clients)}): "
              + ", ".join(f"{c} (${rank[c]:,})" for c in new_clients))
    if gone_clients:
        print(f"NO LONGER PRESENT ({len(gone_clients)}): " + ", ".join(gone_clients))
    if prev is None:
        print("NOTE: could not read the previously published data, so no new-client diff.")
    for note in notes:
        print("NOTE:", note)

    if args.push:
        git = ["git", "-C", REPO]
        if not subprocess.run(git + ["diff", "--quiet", "--", "analytics.html"]).returncode:
            print("No change to analytics.html — nothing to push.")
            return
        subprocess.run(git + ["add", "analytics.html"], check=True)
        subprocess.run(
            git + ["commit", "-m", f"Revenue dashboard: resync from Deal Tracker ({payload['asOf']})"],
            check=True,
        )
        subprocess.run(git + ["push"], check=True)
        print("Pushed to anomalycreativeco.github.io.")


if __name__ == "__main__":
    main()
